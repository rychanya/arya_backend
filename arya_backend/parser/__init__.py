from io import BytesIO

from bson.binary import USER_DEFINED_SUBTYPE
from bson.objectid import ObjectId
from openpyxl import load_workbook

from arya_backend.db import fs
from arya_backend.db.QA import get_or_create_qa_incomplite, is_exists
from arya_backend.db.upload_QA import set_data
from arya_backend.models.qa import QA, QAIncomplete
from arya_backend.models.upload_QA import Foreign, Payload, Upload


def parse_type(type: str) -> str:
    if type == "единственный выбор":
        return QA.type_enum.one.value
    elif type == "множественный выбор":
        return QA.type_enum.many.value
    else:
        raise TypeError


def parse_answer(answer: str, answer_type: str) -> list[str]:
    if answer_type == QA.type_enum.one:
        return [answer]
    elif answer_type == QA.type_enum.many:
        return answer.split(";_x000D_\n")
    else:
        raise TypeError


def parse_correct(value: str) -> bool:
    if value == "+":
        return True
    elif value == "-":
        return False
    else:
        raise TypeError


def normalize_cp1251(s: str) -> str:
    try:
        return s.encode("cp1251").decode("utf8")
    except UnicodeDecodeError:
        return s


# def row_iter(stream: BytesIO, user: ObjectId):
#     wb = load_workbook(filename=stream, read_only=True)
#     for ws in wb.worksheets:
#         title = normalize_cp1251(str(ws.title))
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             try:
#                 type = parse_type(row[6])
#                 answer = parse_answer(row[4], type)
#                 question = row[1]
#                 is_correct = parse_correct(row[2])
#                 yield QAIncomplete.parse_obj(
#                     {
#                         "title": title,
#                         "type": type,
#                         "answer": answer,
#                         "question": question,
#                         "is_correct": is_correct,
#                         "by": [user],
#                         "create": user,
#                     }
#                 )
#             except (TypeError, KeyError):
#                 pass


def row_iter(user_id: ObjectId, payload: list[Payload]):
    for row in payload:
        try:
            title = normalize_cp1251(row.title)
            type = parse_type(row.type)
            answer = parse_answer(row.answer, type)
            question = row.question
            is_correct = parse_correct(row.is_correct)
            yield QAIncomplete.parse_obj(
                {
                    "title": title,
                    "type": type,
                    "answer": answer,
                    "question": question,
                    "is_correct": is_correct,
                    "by": [user_id],
                    "create": user_id,
                }
            )
        except (TypeError, KeyError):
            pass


def parse_qa(qa: QAIncomplete) -> Foreign:
    if qa.is_correct:
        doc_id = is_exists(qa.type, qa.answer, qa.question)
        if doc_id:
            return Foreign(id=doc_id, col="QA")
    return get_or_create_qa_incomplite(qa)


def parse(id: ObjectId, user_id: ObjectId, payload: list[Payload]):
    data = [parse_qa(qa).dict() for qa in row_iter(user_id, payload)]
    set_data(id, data)
